import sqlite3
import json
import os

# コスモジャパン大府店専用エクスポートスクリプト
DB_FILE = 'slot_data_obu.db'
EXCEL_FILE = '島図最新版.xlsx'
EXCEL_SHEET = 'コスモ大府'
DOCS_DIR = os.path.join('docs', 'cosmo_obu')


def export_data():
    os.makedirs(DOCS_DIR, exist_ok=True)

    # 1. DBからJSONへエクスポート
    print(f"Reading from {DB_FILE}...")
    if not os.path.exists(DB_FILE):
        print(f"Warning: {DB_FILE} does not exist. Creating empty data.json.")
        data = []
    else:
        try:
            conn = sqlite3.connect(DB_FILE)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute("SELECT 日付, 機種名, 台番号, BONUS, BIG, REG, 累計ゲーム, 最終差枚 FROM slot_data")
            data = [dict(row) for row in c.fetchall()]
            print(f"Extracted {len(data)} records from DB.")
        except Exception as e:
            print(f"Error reading from DB: {e}")
            data = []
        finally:
            conn.close()

    data_path = os.path.join(DOCS_DIR, 'data.json')
    with open(data_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)
    print(f"Saved {data_path}")

    # 2. Excelレイアウト(コスモ大府シート)からJSONへエクスポート
    print(f"Reading layout from {EXCEL_FILE} sheet='{EXCEL_SHEET}'...")
    if not os.path.exists(EXCEL_FILE):
        print(f"Warning: {EXCEL_FILE} does not exist. Skipping layout.json update.")
    else:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
            if EXCEL_SHEET not in wb.sheetnames:
                print(f"Warning: Sheet '{EXCEL_SHEET}' not found in {EXCEL_FILE}.")
                return
            ws = wb[EXCEL_SHEET]
            layout = []
            for row in ws.iter_rows(values_only=True):
                r = []
                for val in row:
                    if val is None:
                        r.append('')
                    elif isinstance(val, float) and val.is_integer():
                        r.append(int(val))
                    elif isinstance(val, (int, float)):
                        r.append(int(val))
                    else:
                        r.append(str(val))
                layout.append(r)
            # 末尾の全空行を除去
            while layout and all(c == '' for c in layout[-1]):
                layout.pop()
            print(f"Extracted layout with {len(layout)} rows (after trimming empty rows).")
            layout_path = os.path.join(DOCS_DIR, 'layout.json')
            with open(layout_path, 'w', encoding='utf-8') as f:
                json.dump(layout, f, ensure_ascii=False)
            print(f"Saved {layout_path}")
        except ImportError:
            print("openpyxl not found. Skipping layout.json update.")
        except Exception as e:
            print(f"Error reading Excel layout: {e}")


if __name__ == "__main__":
    export_data()
    print("Export complete!")
