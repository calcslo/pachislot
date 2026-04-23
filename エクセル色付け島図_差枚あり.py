import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Excelファイルの読み込み
file_path = "modifiedrowdata有楽砂川.xlsx"  # Excelファイルのパス
wb = load_workbook(file_path)

# 元データを読み込み（イベント日1、イベント日2、イベント日3、全体集計）
def load_sheet_data(sheet_name):
    df = pd.DataFrame(wb[sheet_name].values, columns=["台番号", "差枚合計"])
    df["差枚合計"] = pd.to_numeric(df["差枚合計"], errors="coerce")
    return df

event_data = {
    "全体集計": load_sheet_data("全体集計"),
    "イベント日1": load_sheet_data("イベント日1集計"),
    "イベント日2": load_sheet_data("イベント日2集計"),
    "イベント日3": load_sheet_data("イベント日3集計"),
}

# 範囲と色の指定（プラスとマイナスで色分け）
# プラスとマイナスの範囲を計算
def calculate_ranges(df):
    # プラスとマイナスの最小値・最大値を計算
    min_value = df["差枚合計"].min()
    max_value = df["差枚合計"].max()

    # 0を基準に分割
    plus_min = 0
    plus_max = max_value
    minus_min = min_value
    minus_max = 0

    # プラス、マイナスそれぞれを2分割する
    plus_range = (plus_max - plus_min) / 2
    minus_range = (minus_max - minus_min) / 2

    return plus_min, plus_max, plus_range, minus_min, minus_max, minus_range

# 色を決定する関数
def get_color(value, plus_min, plus_max, plus_range, minus_min, minus_max, minus_range):
    # プラスとマイナスの色を分けて判定
    if value >= plus_min:
        if value > (plus_max - plus_range):
            return "FF0000"  # 赤
        else:
            return "32CD32"  # 緑
    else:
        if value < (minus_min + minus_range):
            return "0000FF"  # 青
        else:
            return "00BFFF"  # 水色

# 島図シートのコピーまたは上書き
original_sheet = wb["有楽砂川"]
sheets = {}
for name in event_data.keys():
    sheet_name = f"{name}集計島図"
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # 既存のシートをクリア
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.value = None
    else:
        sheet = wb.copy_worksheet(original_sheet)
        sheet.title = sheet_name
    sheets[name] = sheet

# 色分けと基準の書き込みを適用
for name, df_data in event_data.items():
    sheet = sheets[name]
    plus_min, plus_max, plus_range, minus_min, minus_max, minus_range = calculate_ranges(df_data)

    # 色分け基準をシートに書き込む
    sheet["M1"] = "基準"
    sheet["N1"] = f"範囲: {plus_range:.0f} ～ {plus_max:.0f} (赤)"
    sheet["N2"] = f"範囲: {plus_min:.0f} ～ {plus_range:.0f} (緑)"
    sheet["N3"] = f"範囲: -{minus_range:.0f} ～ {minus_max:.0f} (水色)"
    sheet["N4"] = f"範囲: {minus_min:.0f} ～ -{minus_range:.0f} (青色)"


    # 色分けの適用
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value in df_data["台番号"].values:
                # 差枚の値を取得
                diff = df_data.loc[df_data["台番号"] == cell.value, "差枚合計"].values[0]
                if pd.notna(diff):  # NaN チェック
                    color = get_color(diff, plus_min, plus_max, plus_range, minus_min, minus_max, minus_range)
                    if color:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# 保存
wb.save(file_path)
print("プラスとマイナスの範囲に応じた色分けが完了しました！")
