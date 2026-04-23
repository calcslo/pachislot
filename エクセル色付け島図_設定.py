import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Excelファイルの読み込み
file_path = "modifiedrowdataogiyatest1.xlsx"  # Excelファイルのパス
wb = load_workbook(file_path)

# 元データを読み込み（イベント日1、イベント日2、イベント日3、全体集計）
def load_sheet_data(sheet_name):
    df = pd.DataFrame(wb[sheet_name].values, columns=["台番号","機種名","G数合計","BB回数合計","RB回数合計", "設定合計"])
    df["設定合計"] = pd.to_numeric(df["設定合計"], errors="coerce")
    return df

event_data = {
    "全体集計": load_sheet_data("全体集計1"),
    "イベント日1": load_sheet_data("イベント日1集計1"),
    "イベント日2": load_sheet_data("イベント日2集計1"),
    "イベント日3": load_sheet_data("イベント日3集計1"),
}

# 設定に応じた色を決定する関数
def get_color_by_setting(value):
    color_map = {
        1: "FFFFFF",  # 白
        2: "0000FF",  # 青
        3: "FFFF00",  # 黄色
        4: "32CD32",  # 緑
        5: "FF0000",  # 赤
        6: "800080",  # 紫
    }
    return color_map.get(value, None)

# 島図シートのコピーまたは上書き
original_sheet = wb["オーギヤ半田"]
sheets = {}
for name in event_data.keys():
    sheet_name = f"{name}集計島図設定版"
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # 既存のシートをクリア
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.value = None
    else:
        sheet = wb.copy_worksheet(original_sheet)
        sheet.title = sheet_name
    sheets[name] = sheet

# 色分けの適用
for name, df_data in event_data.items():
    sheet = sheets[name]

    # 色分けの適用
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value in df_data["台番号"].values:
                # 設定の値を取得
                setting = df_data.loc[df_data["台番号"] == cell.value, "設定合計"].values[0]
                if pd.notna(setting):  # NaN チェック
                    color = get_color_by_setting(setting)
                    if color:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# 保存
wb.save(file_path)
print("設定1～6に応じた色分けが完了しました！")
