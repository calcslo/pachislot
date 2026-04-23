import openpyxl
import pandas as pd

# ファイルパスを指定
file_path = 'modifiedrowdataogiyatest1.xlsx'

# Excelファイルを読み込む
wb = openpyxl.load_workbook(file_path)

# アクティブシートを取得
ws = wb["modifiedrowdataogiyatest1"]

# データをPandas DataFrameに読み込む
data = []
for row in ws.iter_rows(min_row=2, values_only=True):  # ヘッダーをスキップするためmin_row=2
    台番号 = row[1]  # B列はインデックス1
    予測差枚 = row[23]  # X列はインデックス23
    イベント日1 = row[18]  # Y列はインデックス24
    イベント日2 = row[19]  # Z列はインデックス25
    イベント日3 = row[20]  # AA列はインデックス26
    data.append([台番号, 予測差枚, イベント日1, イベント日2, イベント日3])

# DataFrameに変換
df = pd.DataFrame(data, columns=['台番号', '予測差枚', 'イベント日1', 'イベント日2', 'イベント日3'])

# 台番号ごとに全体を集計
aggregated_data_all = df.groupby('台番号')['予測差枚'].sum().reset_index()

# イベント日1, 2, 3が1の場合のデータをそれぞれフィルタリングして集計
aggregated_data_event1 = df[df['イベント日1'] == 1].groupby('台番号')['予測差枚'].sum().reset_index()
aggregated_data_event2 = df[df['イベント日2'] == 1].groupby('台番号')['予測差枚'].sum().reset_index()
aggregated_data_event3 = df[df['イベント日3'] == 1].groupby('台番号')['予測差枚'].sum().reset_index()

# 新しいシートを作成
ws_all = wb.create_sheet('全体集計')
ws_event1 = wb.create_sheet('イベント日1集計')
ws_event2 = wb.create_sheet('イベント日2集計')
ws_event3 = wb.create_sheet('イベント日3集計')

# 各集計結果を新しいシートに書き込む
def write_to_sheet(sheet, data, header):
    sheet.append(header)  # ヘッダーを追加
    for _, row in data.iterrows():
        sheet.append(row.tolist())

# 全体集計を出力
write_to_sheet(ws_all, aggregated_data_all, ['台番号', '予測差枚合計'])

# イベント日1, 2, 3の集計を出力
write_to_sheet(ws_event1, aggregated_data_event1, ['台番号', '予測差枚合計'])
write_to_sheet(ws_event2, aggregated_data_event2, ['台番号', '予測差枚合計'])
write_to_sheet(ws_event3, aggregated_data_event3, ['台番号', '予測差枚合計'])

# 変更を保存
wb.save('modifiedrowdataogiyatest1.xlsx')