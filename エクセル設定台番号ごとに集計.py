import openpyxl
import pandas as pd
from 設定判別ほんとに累計版 import cal_setting_short

# ファイルパスを指定
file_path = 'modifiedrowdataogiyatest1.xlsx'

# Excelファイルを読み込む
wb = openpyxl.load_workbook(file_path)

# アクティブシートを取得
ws = wb["modifiedrowdataogiyatest1"]

# データを読み込み
data = []
for row in ws.iter_rows(min_row=2, values_only=True):  # ヘッダーをスキップするためmin_row=2
    機種名 = row[0]  # B列はインデックス1
    台番号 = row[1]  # C列はインデックス2
    G数 = row[2]  # D列はインデックス3
    BB回数 = row[3]  # E列はインデックス4
    RB回数 = row[4]  # F列はインデックス5
    イベント日1 = row[16]  # Y列はインデックス24
    イベント日2 = row[17]  # Z列はインデックス25
    イベント日3 = row[18]  # AA列はインデックス26

    data.append([機種名, 台番号, G数, BB回数, RB回数, イベント日1, イベント日2, イベント日3])

# DataFrameに変換
df = pd.DataFrame(data, columns=['機種名', '台番号', 'G数', "BB回数", 'RB回数', 'イベント日1', 'イベント日2', 'イベント日3'])

# 台番号ごとに全体を集計
aggregated_data = df.groupby(['台番号', '機種名']).agg({'G数': 'sum', 'BB回数': 'sum', 'RB回数': 'sum'}).reset_index()

# 予測設定を追加
aggregated_data['予測設定'] = aggregated_data.apply(
    lambda row: cal_setting_short(row['BB回数'], row['RB回数'], row['G数'], row['機種名']),
    axis=1
)

# 集計結果をExcelに出力
def write_to_sheet_with_prediction(sheet, data, header):
    sheet.append(header)  # ヘッダーを追加
    for _, row in data.iterrows():
        sheet.append(row.tolist())

# 全体集計シートを作成
ws_all = wb.create_sheet('全体集計')
write_to_sheet_with_prediction(
    ws_all,
    aggregated_data,
    ['台番号', '機種名', 'G数合計', 'BB回数合計', 'RB回数合計', '予測設定']
)

# イベント日ごとに集計してシート作成
for event_col, sheet_name in zip(['イベント日1', 'イベント日2', 'イベント日3'], ['イベント日1集計', 'イベント日2集計', 'イベント日3集計']):
    filtered_data = df[df[event_col] == 1].groupby(['台番号', '機種名']).agg({'G数': 'sum', 'BB回数': 'sum', 'RB回数': 'sum'}).reset_index()
    filtered_data['予測設定'] = filtered_data.apply(
        lambda row: cal_setting_short(row['BB回数'], row['RB回数'], row['G数'], row['機種名']),
        axis=1
    )
    ws_event = wb.create_sheet(sheet_name)
    write_to_sheet_with_prediction(
        ws_event,
        filtered_data,
        ['台番号', '機種名', 'G数合計', 'BB回数合計', 'RB回数合計', '予測設定']
    )

# 変更を保存
wb.save('modifiedrowdataogiyatest1.xlsx')
