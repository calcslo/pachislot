import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import matplotlib.pyplot as plt
from tempfile import NamedTemporaryFile
import numpy as np
from scipy import stats
import seaborn as sns
from scipy.stats import spearmanr

#sp = "オーギヤ半田島図.xlsx"
#op = "オーギヤ半田色付け島図画像.png
class MyClass:
    def __init__(self):
        plt.rcParams["font.family"] = "MS Gothic"  # Windows用

        pass
    def shimazu(self, df, shimazu_path, output_path):
        plt.rcParams["font.family"] = "MS Gothic"  # Windows用
        grouped = df.groupby("台番号")["予測差枚"].mean().reset_index()
        min_value = grouped["予測差枚"].min()
        #print(min_value)
        max_value = grouped["予測差枚"].max()
        #print(max_value)
        self.plus_max = max_value
        self.minus_min = min_value

        # プラス、マイナスそれぞれを2分割する
        self.plus_range = self.plus_max / 3
        self.minus_range = -self.minus_min / 2


        wb = load_workbook(shimazu_path)
        ws = wb["島図"]
        max_col = 24  # AG列
        max_row = 44  # 50行目

        # 色塗り処理（指定範囲内のみ）
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=3, max_col=max_col):
            for cell in row:
                if cell.value in grouped["台番号"].values:
                    diff = grouped[grouped["台番号"] == cell.value]["予測差枚"].values[0]
                    color_code = self.get_color(diff)
                    if color_code:
                        fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
                        cell.fill = fill


        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpfile:
            wb.save(tmpfile.name)


            # PILで画像化（Excelの画面キャプチャは直接できないので、代替処理）
            fig_width = max_col * 1
            fig_height = max_row * 0.25
            fig, ax = plt.subplots(figsize=(fig_width, fig_height))
            ax.axis("off")
            ax.set_title("ハナハナ　差枚　データ\n"
                         f"青:～{int(self.minus_min + self.minus_range)} / 水色:{int(self.minus_min + self.minus_range)}～-1 / 黄色:1～{int(self.plus_max - 2*self.plus_range)} / 緑:{int(self.plus_max - 2*self.plus_range)}～{int(self.plus_max - self.plus_range)} / 赤:{int(self.plus_max - self.plus_range)}～", fontsize=20)

            # セル内容をmatplotlibに描画
            for i, row in enumerate(ws.iter_rows(max_row=max_row, min_col=3, max_col=max_col)):
                for j, cell in enumerate(row):
                    val = str(cell.value) if cell.value is not None else ""
                    fill_color = cell.fill.start_color.rgb[-6:] if cell.fill.fill_type == "solid" else "FFFFFF"
                    rect = plt.Rectangle((j, -i), 1, 1, facecolor=f"#{fill_color}", edgecolor="black")
                    ax.add_patch(rect)
                    ax.text(j + 0.5, -i + 0.5, val, va='center', ha='center', fontsize=11)

            ax.set_xlim(0, max_col-2)
            ax.set_ylim(-max_row + 1, 0)  # Y軸の範囲を調整して全行が描画されるように
            plt.subplots_adjust(top=0.9)  # タイトルが格子と被らないように余白を追加
            plt.savefig(output_path, dpi=300, bbox_inches='tight')
            plt.close()
    def get_color(self, diff):
        if diff > 0:
            if diff > (self.plus_max - self.plus_range):
                return "F08080"  # 赤
            elif diff > (self.plus_max - 2*self.plus_range):
                return "90EE90"  # 緑
            else:
                return "FFF797"  # 黄色
        elif diff < 0:
            if diff < (self.minus_min + self.minus_range):
                return "749DFF"  # 青
            else:
                return "B0E0E6"  # 水色
        elif diff == 0:
            return "FFFFFF"
        else:
            return None

    def grid_setting6(self, df, shimazu_path, output_path):
        plt.rcParams["font.family"] = "MS Gothic"  # Windows用
        wb = load_workbook(shimazu_path)
        ws = wb["島図"]
        max_col = 24  # AG列
        max_row = 44  # 50行目
        tai_bango_dict = {}
        # CSVの「台番号」と「角からの位置」を基にデータを取得
        for index, row in df.iterrows():
            tai_bango = row["台番号"]
            kaku_kara_no_ichi = row["推定高設定"]
            tai_bango_dict[tai_bango] = kaku_kara_no_ichi

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpfile:
            wb.save(tmpfile.name)

            # PILで画像化（Excelの画面キャプチャは直接できないので、代替処理）
            fig_width = max_col * 1
            fig_height = max_row * 0.25
            fig, ax = plt.subplots(figsize=(fig_width, fig_height))
            ax.axis("off")

            # タイトルの位置調整 (余白を追加)
            ax.set_title("推　定　高　設　定　の　箇　所", fontsize=20)

            # セル内容をmatplotlibに描画
            for i, row in enumerate(ws.iter_rows(max_row=max_row, min_col=3,max_col=max_col)):
                for j, cell in enumerate(row):
                    val = str(cell.value) if cell.value is not None else ""
                    fill_color = cell.fill.start_color.rgb[-6:] if cell.fill.fill_type == "solid" else "FFFFFF"
                    rect = plt.Rectangle((j, -i), 1, 1, facecolor=f"#{fill_color}", edgecolor="black")
                    ax.add_patch(rect)

                    ax.text(j + 0.5, -i + 0.5, val, va='center', ha='center', fontsize=11)
            for i, row in enumerate(ws.iter_rows(max_row=max_row, min_col=3,max_col=max_col)):
                for j, cell in enumerate(row):
                    tai_bango = cell.value
                    if tai_bango in tai_bango_dict and tai_bango_dict[tai_bango] == 1:
                        rect = plt.Rectangle((j, -i), 1, 1, facecolor="none", edgecolor="red", linewidth=3)  # 赤色で太い枠線
                        ax.add_patch(rect)

            # Y軸の調整: 最初の行が画像内に収まるように調整
            ax.set_xlim(0, max_col-2)
            ax.set_ylim(-max_row +1, 0)  # Y軸の範囲調整

            # 余白の調整
            plt.subplots_adjust(top=0.9)  # タイトルが格子と被らないように余白を追加

            plt.savefig(output_path, dpi=300, bbox_inches='tight')
            plt.close()

    def plot_daily_samai(self, df, graph_path):
        df["日別差枚平均"] = df["日別差枚平均"].apply(pd.to_numeric, errors='coerce')
        # 日別平均差枚を基にした横棒グラフ
        fig, ax = plt.subplots(figsize=(10, 35.5))
        y_positions = np.arange(len(df))  # y軸の位置
        colors = ['red' if x < 0 else 'blue' for x in df['日別差枚平均']]

        df['日付'] = pd.to_datetime(df['日付'], errors='coerce')

        ax.barh(df['日付'].dt.strftime('%Y-%m-%d'), df['日別差枚平均'], color=colors)
        for i in range(len(y_positions)):
            if i % 2 == 0:
                ax.axhspan(i - 0.5, i + 0.5, facecolor='lightgray', alpha=0.3)  # 偶数行に灰色背景

        # 0を基準にする
        ax.axvline(x=0, color='gray', linestyle='--')

        # グラフのタイトルとラベルを設定
        ax.set_xlabel('日別平均差枚')
        ax.set_title('')
        # 横軸メモリを上と下両方に表示
        ax.xaxis.set_ticks_position('bottom')
        ax.xaxis.set_ticks_position('top')
        # 縦軸を逆転
        ax.invert_yaxis()
        ax.margins(y=0)  # y軸のマージンを0に設定

        plt.subplots_adjust(top=0.95, bottom=0)
        plt.tight_layout()

        # 画像を保存
        plt.savefig(graph_path)
    def x_day(self, df, graph_path):
        average_diff_by_x = df.groupby("xのつく日")["予測差枚"].mean()

        # 2. 各値を個別の変数に代入（例：x0, x1, ..., x9）
        x0, x1, x2, x3, x4, x5, x6, x7, x8, x9 = [average_diff_by_x.get(i, 0) for i in range(10)]

        x_labels = list(range(10))
        x_values = [average_diff_by_x.get(i, 0) for i in x_labels]

        # 色分け（正：青、負：赤）
        colors = ['blue' if val >= 0 else 'red' for val in x_values]

        # 横棒グラフ描画
        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.barh(x_labels, x_values, color=colors)

        # y軸ラベルを「0のつく日」〜「9のつく日」
        ax.set_yticks(x_labels)
        ax.set_yticklabels([f"{i}のつく日" for i in x_labels])
        ax.set_xlabel("平均差枚")
        ax.set_title("xのつく日の平均差枚")

        # 値を横に表示
        for bar in bars:
            width = bar.get_width()
            ax.text(width + (1 if width >= 0 else -1), bar.get_y() + bar.get_height() / 2,
                    f"{width:.0f}", ha='left' if width >= 0 else 'right', va='center')

        plt.tight_layout()
        plt.savefig(graph_path)
        return x0, x1, x2, x3, x4, x5, x6, x7, x8, x9
    def weekday(self,df,graph_path):

        # 仮のDataFrame（実際にはdfに「曜日」と「予測差枚」列がある前提）
        # df = pd.DataFrame({
        #     "曜日": [...],  # 例: "月", "火", "水", ...
        #     "予測差枚": [...]
        # })

        # 曜日順の定義
        week_order = ["月", "火", "水", "木", "金", "土", "日"]

        # 平均予測差枚を曜日で集計
        weekday_means = df.groupby("曜日")["予測差枚"].mean()
        w0 = weekday_means["月"]
        w1 = weekday_means["火"]
        w2 = weekday_means["水"]
        w3 = weekday_means["木"]
        w4 = weekday_means["金"]
        w5 = weekday_means["土"]
        w6 = weekday_means["日"]

        # 指定順で揃える（ない曜日は0で埋める）
        y_labels = week_order
        y_values = [weekday_means.get(day, 0) for day in y_labels]

        # 色分け（正：青、負：赤）
        colors = ['blue' if val >= 0 else 'red' for val in y_values]

        # 横棒グラフ描画
        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.barh(y_labels, y_values, color=colors)

        ax.set_xlabel("平均差枚")
        ax.set_title("曜日ごとの平均差枚")

        # 値を棒の横に表示
        for bar in bars:
            width = bar.get_width()
            ax.text(width + (1 if width >= 0 else -1), bar.get_y() + bar.get_height() / 2,
                    f"{width:.0f}", ha='left' if width >= 0 else 'right', va='center')

        plt.tight_layout()
        plt.savefig(graph_path)
        plt.close()
        return w0, w1, w2, w3, w4, w5, w6

    def out_ratio(self, total_game, samai):
        in_medals = total_game * 3
        out_medals = samai + in_medals

        out_ratio = out_medals / in_medals
        return round(out_ratio, 3)

    def matubi_kentei(self, df):
        results_df = pd.DataFrame(columns=['日付'] + [f'有意差末尾{i}' for i in range(10)] + ["末尾寄せ"])
        # 各日付ごとに処理
        for date, group in df.groupby('日付'):
            # 日付ごとのすべての台番号データを抽出
            all_data = group['予測差枚']

            # 末尾（0～9）ごとにデータを抽出し、有意差を判定
            significant_tails = {i: 0 for i in range(10)}  # 最初はすべての末尾に有意差なし（0）
            tail_list = []
            for tail in range(10):
                # 特定の末尾番号に該当する台番号のデータを抽出
                tail_data = group[group['台番号末尾'] == tail]['予測差枚']

                if len(tail_data) > 1 and len(all_data) > 1:
                    # t検定を実行
                    t_stat, p_value = stats.ttest_ind(all_data, tail_data)

                    # 有意水準5%で有意差を判定
                    if p_value < 0.05:
                        # 有意差があった場合、該当する末尾に1をセット
                        significant_tails[tail] = 1
                        tail_list.append(tail)

                    # 日付ごとの結果をresults_dfに追加
            row = {'日付': date}
            row.update({f'末尾{i}': significant_tails[i] for i in range(10)})

            # 末尾寄せの列に有意差のある末尾を文字列として追加
            row["末尾寄せ"] = ",".join(map(str, tail_list)) if tail_list else "なし"

            results_df = pd.concat([results_df, pd.DataFrame([row])], ignore_index=True)

        results_df = results_df[['日付', '末尾寄せ']]
        return results_df

    def hanyo_shukei(self, df, target, graph_path):
        # ターゲットごとの平均差枚を計算
        target_means = df.groupby(target)["予測差枚"].mean().reset_index()

        # グラフのサイズ設定
        fig, ax = plt.subplots(figsize=(10, 6))

        # 各バーの色を設定（予測差枚が負なら赤、正なら青）
        colors = ['red' if x < 0 else 'blue' for x in target_means["予測差枚"]]

        # 水平棒グラフを作成
        bars = ax.barh(target_means[target], target_means["予測差枚"], color=colors)

        for i in range(len(target_means)):
            if i % 2 == 0:
                ax.axhspan(i - 0.5, i + 0.5, facecolor='lightgray', alpha=0.3)  # 偶数行に灰色背景

        # 軸ラベルとタイトル設定
        ax.set_xlabel('平均差枚', fontsize=14)
        ax.set_title(f'{target}ごとの平均差枚', fontsize=16)

        # 値を棒の横に表示
        for bar in bars:
            width = bar.get_width()
            ax.text(width + (1 if width >= 0 else -1), bar.get_y() + bar.get_height() / 2,
                    f"{width:.0f}", ha='left' if width >= 0 else 'right', va='center')
        # y軸のラベルを逆順に設定
        ax.invert_yaxis()

        # y軸のラベルを省略せずに表示
        ax.set_yticks(range(len(target_means[target])))  # y軸の位置を設定
        ax.set_yticklabels(target_means[target], fontsize=12)

        # グラフのレイアウトを調整
        plt.tight_layout()

        # グラフを指定されたパスに保存
        plt.savefig(graph_path)

        # 描画を終了
        plt.close()


    def combine_files(self, old_path, new_path, file_name):

        # file1の拡張子をチェックして読み込み方を決定
        df1 = pd.read_csv(old_path)
        df2 = pd.read_csv(new_path)

        # df1とdf2を縦に結合する
        df_combined = pd.concat([df1, df2], axis=0).drop_duplicates()
        df_combined = df_combined.drop_duplicates()
        # ファイルを保存する
        df_combined.to_csv(file_name, index=False)

        print("縦に結合されたファイルが保存されました:", file_name)
        return file_name

    def yuui_kentei(self, df, x_col, y_col, output_path):
        """
           指定されたDataFrameの2列に対してスピアマン相関と近似曲線をプロット。

           Parameters:
               df (pd.DataFrame): 入力データ
               x_col (str): x軸の列名
               y_col (str): y軸の列名
               degree (int): 近似式の次数（既定は2次）
           """
        degree = 1
        # データの前処理（NaN除去）
        data = df[[x_col, y_col]].dropna()

        # スピアマン相関
        corr, p_value = spearmanr(data[x_col], data[y_col])
        print(f"スピアマン相関係数: {corr:.4f}（p = {p_value:.4g}）")

        # 散布図と近似曲線の描画
        plt.figure(figsize=(8, 6))
        sns.scatterplot(x=x_col, y=y_col, data=data, label='データ点')

        # 近似式の計算（多項式フィッティング）
        x = data[x_col].values
        y = data[y_col].values
        coeffs = np.polyfit(x, y, degree)
        poly = np.poly1d(coeffs)
        x_fit = np.linspace(x.min(), x.max(), 100)
        y_fit = poly(x_fit)
        plt.plot(x_fit, y_fit, color='red', label=f'{degree}次近似曲線')

        # 式を描画
        eq_str = " + ".join([f"{c:.3g}·x^{i}" if i > 0 else f"{c:.3g}" for i, c in enumerate(coeffs[::-1])])
        plt.text(0.05, 0.95, f"近似式: y = {eq_str}", transform=plt.gca().transAxes,
                 fontsize=10, verticalalignment='top', bbox=dict(boxstyle="round", fc="w"))

        plt.title(f"{x_col} vs {y_col} のスピアマン相関と近似曲線")
        plt.legend()
        plt.grid(True)
        plt.tight_layout()
        plt.savefig(output_path)
        #plt.show()
        plt.close()
"""
mod = MyClass()
df = pd.read_csv("3monthsrowdata上小田井.csv", encoding="utf-8")
df = df[df["イベント日"] == 1]
df = df[df["機種名"] == "キングハナハナ-30"]
x_col = "前日RB回数"
y_col = "予測差枚"
output_path = "相関係数テスト.png"
mod.yuui_kentei(df, x_col, y_col, output_path)
"""